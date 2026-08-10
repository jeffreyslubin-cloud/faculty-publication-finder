from __future__ import annotations

import re
import time
from pathlib import Path
from typing import Any

import pandas as pd
from Bio import Entrez


START_DATE = "2025/07/01"
END_DATE = "2026/06/30"

ROSTER_FILE = "Penn_State_EM_Faculty.xlsx"
CLEANED_FILE = "FY26 Publications.xlsx"
OUTPUT_FILE = "FY26 Supplemental Candidates.xlsx"

Entrez.email = "jeffrey.s.lubin@gmail.com"


def clean(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return re.sub(r"\s+", " ", str(value)).strip()


def faculty_name(row: pd.Series) -> str:
    last = clean(row.get("Last Name"))
    first = clean(row.get("First Name / Initial"))
    return f"{last}, {first}" if first else last


def build_query(row: pd.Series) -> str:
    """
    Search by PubMed entry date (EDAT), not article publication date.

    Important: the Author search is intentionally NOT quoted. PubMed then
    automatically truncates author initials, so Smith J[Author] can retrieve
    Smith JA, Smith JB, etc. This improves sensitivity.
    """
    last = clean(row.get("Last Name"))
    first = clean(row.get("First Name / Initial"))
    initial = clean(row.get("PubMed Initials"))[:1]

    author_terms = [f"{last} {initial}[Author]"]

    # Add the exact full-name form as a second route when a full first name exists.
    if len(first) > 1:
        author_terms.append(f'"{first} {last}"[Full Author Name]')

    return (
        f'({" OR ".join(author_terms)}) '
        f'AND {START_DATE}:{END_DATE}[EDAT]'
    )


def search_pmids(query: str) -> list[str]:
    with Entrez.esearch(
        db="pubmed",
        term=query,
        retmax=10000,
        sort="pub date",
    ) as handle:
        result = Entrez.read(handle)
    return [str(x) for x in result.get("IdList", [])]


def fetch_records(pmids: list[str]) -> list[dict]:
    records: list[dict] = []
    for start in range(0, len(pmids), 200):
        batch = pmids[start:start + 200]
        with Entrez.efetch(
            db="pubmed",
            id=",".join(batch),
            rettype="medline",
            retmode="xml",
        ) as handle:
            data = Entrez.read(handle)
        records.extend(data.get("PubmedArticle", []))
        time.sleep(0.4)
    return records


def history_date(record: dict, wanted_status: str) -> str:
    for item in record.get("PubmedData", {}).get("History", []):
        status = clean(getattr(item, "attributes", {}).get("PubStatus")).lower()
        if status == wanted_status.lower():
            year = clean(item.get("Year"))
            month = clean(item.get("Month")).zfill(2)
            day = clean(item.get("Day")).zfill(2)
            return "-".join(x for x in [year, month, day] if x)
    return ""


def publication_date(record: dict) -> str:
    pub_date = (
        record.get("MedlineCitation", {})
        .get("Article", {})
        .get("Journal", {})
        .get("JournalIssue", {})
        .get("PubDate", {})
    )

    year = clean(pub_date.get("Year"))
    month = clean(pub_date.get("Month"))
    day = clean(pub_date.get("Day"))
    medline_date = clean(pub_date.get("MedlineDate"))

    if year:
        return "-".join(x for x in [year, month, day] if x)
    return medline_date


def parse_record(record: dict) -> dict[str, Any]:
    citation = record.get("MedlineCitation", {})
    article = citation.get("Article", {})
    journal = article.get("Journal", {})

    author_details: list[dict[str, str]] = []
    author_display: list[str] = []
    affiliations: list[str] = []

    for author in article.get("AuthorList", []):
        collective = clean(author.get("CollectiveName"))
        if collective:
            author_display.append(collective)
            continue

        last = clean(author.get("LastName"))
        fore = clean(author.get("ForeName"))
        initials = clean(author.get("Initials"))

        author_details.append(
            {"last": last, "fore": fore, "initials": initials}
        )
        author_display.append(", ".join(x for x in [last, fore] if x))

        for item in author.get("AffiliationInfo", []):
            affiliation = clean(item.get("Affiliation"))
            if affiliation:
                affiliations.append(affiliation)

    doi = ""
    for article_id in record.get("PubmedData", {}).get("ArticleIdList", []):
        if getattr(article_id, "attributes", {}).get("IdType") == "doi":
            doi = clean(article_id)
            break

    return {
        "PMID": clean(citation.get("PMID")),
        "Title": clean(article.get("ArticleTitle")),
        "Journal": clean(journal.get("Title")),
        "Publication Date": publication_date(record),
        "PubMed Entry Date": (
            history_date(record, "entrez")
            or history_date(record, "pubmed")
        ),
        "Authors": "; ".join(author_display),
        "Affiliations": " | ".join(dict.fromkeys(affiliations)),
        "DOI": doi,
        "_author_details": author_details,
    }


def classify(row: pd.Series, article: dict[str, Any]) -> tuple[str, str]:
    last = clean(row.get("Last Name")).lower()
    first = clean(row.get("First Name / Initial")).lower()
    initial = clean(row.get("PubMed Initials"))[:1].lower()

    matches = [
        author for author in article["_author_details"]
        if author["last"].lower() == last
    ]

    if not matches:
        return "Unlikely", "No exact surname match in author list"

    affiliation_text = article["Affiliations"].lower()
    penn_state = any(
        term in affiliation_text
        for term in [
            "penn state",
            "pennsylvania state university",
            "hershey medical center",
            "penn state health",
        ]
    )

    secondary = clean(row.get("Secondary Institution(s)"))
    secondary_match = False
    if secondary:
        for part in re.split(r"[;,|]", secondary):
            part = part.strip().lower()
            if part and part in affiliation_text:
                secondary_match = True
                break

    if len(first) > 1:
        for author in matches:
            fore = author["fore"].lower()
            if (
                fore == first
                or fore.startswith(first + " ")
                or first.startswith(fore + " ")
            ):
                reason = "Exact given-name and surname match"
                if penn_state:
                    reason += "; Penn State/Hershey affiliation"
                elif secondary_match:
                    reason += "; secondary affiliation"
                return "Likely", reason

        for author in matches:
            if author["initials"].lower().startswith(initial):
                if penn_state:
                    return "Likely", "Surname/initial match; Penn State/Hershey affiliation"
                if secondary_match:
                    return "Likely", "Surname/initial match; secondary affiliation"
                return "Needs Review", "Surname and first initial match"

        return "Needs Review", "Surname matches, but given name/initials differ"

    for author in matches:
        if author["initials"].lower().startswith(initial):
            if penn_state:
                return "Likely", "Initial/surname match; Penn State/Hershey affiliation"
            if secondary_match:
                return "Likely", "Initial/surname match; secondary affiliation"
            return "Needs Review", "Roster contains only a first initial"

    return "Needs Review", "Surname matches, but initials differ"


def shorten_for_excel(value: Any, limit: int = 32000) -> str:
    text = clean(value)
    if len(text) <= limit:
        return text
    return text[: limit - 24] + " ... [truncated]"


def main() -> None:
    folder = Path(__file__).resolve().parent
    roster_path = folder / ROSTER_FILE
    cleaned_path = folder / CLEANED_FILE
    output_path = folder / OUTPUT_FILE

    if not roster_path.exists():
        raise FileNotFoundError(f"Missing roster: {roster_path}")
    if not cleaned_path.exists():
        raise FileNotFoundError(f"Missing cleaned workbook: {cleaned_path}")

    roster = pd.read_excel(roster_path, sheet_name="Faculty Roster")
    cleaned = pd.read_excel(cleaned_path, sheet_name=0)

    if "PMID" not in cleaned.columns:
        raise ValueError(
            f"{CLEANED_FILE} must contain a PMID column in its first sheet."
        )

    kept_pmids = {
        clean(value)
        for value in cleaned["PMID"].tolist()
        if clean(value)
    }

    print(f"PMIDs already kept: {len(kept_pmids)}")
    print(f"Searching by PubMed Entry Date: {START_DATE} through {END_DATE}\n")

    article_cache: dict[str, dict[str, Any]] = {}
    rows: list[dict[str, Any]] = []

    valid_roster = roster.dropna(subset=["Last Name", "PubMed Initials"]).copy()

    for number, (_, person) in enumerate(valid_roster.iterrows(), start=1):
        name = faculty_name(person)
        query = build_query(person)

        print(f"[{number}/{len(valid_roster)}] {name}")
        found = search_pmids(query)
        supplemental = [pmid for pmid in found if pmid not in kept_pmids]
        print(
            f"    Retrieved: {len(found)} | "
            f"Already kept: {len(found) - len(supplemental)} | "
            f"New candidates: {len(supplemental)}"
        )

        missing = [pmid for pmid in supplemental if pmid not in article_cache]
        if missing:
            for raw in fetch_records(missing):
                parsed = parse_record(raw)
                if parsed["PMID"]:
                    article_cache[parsed["PMID"]] = parsed

        for pmid in supplemental:
            article = article_cache.get(pmid)
            if not article:
                continue

            confidence, reason = classify(person, article)

            rows.append(
                {
                    "Faculty Name": name,
                    "Confidence": confidence,
                    "Reason": reason,
                    "PMID": article["PMID"],
                    "PubMed Entry Date": article["PubMed Entry Date"],
                    "Publication Date": article["Publication Date"],
                    "Title": article["Title"],
                    "Journal": article["Journal"],
                    "Authors": article["Authors"],
                    "Affiliations": shorten_for_excel(article["Affiliations"]),
                    "DOI": article["DOI"],
                    "Search Query": query,
                }
            )

        time.sleep(0.4)

    columns = [
        "Faculty Name",
        "Confidence",
        "Reason",
        "PMID",
        "PubMed Entry Date",
        "Publication Date",
        "Title",
        "Journal",
        "Authors",
        "Affiliations",
        "DOI",
        "Search Query",
    ]
    results = pd.DataFrame(rows, columns=columns)

    if not results.empty:
        rank = {"Likely": 0, "Needs Review": 1, "Unlikely": 2}
        results["_rank"] = results["Confidence"].map(rank).fillna(9)
        results = (
            results.sort_values(
                ["_rank", "Faculty Name", "PubMed Entry Date", "PMID"],
                ascending=[True, True, False, True],
            )
            .drop(columns=["_rank"])
            .drop_duplicates(subset=["Faculty Name", "PMID"])
        )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        results.to_excel(writer, sheet_name="New Candidates Only", index=False)

        worksheet = writer.book["New Candidates Only"]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        widths = {
            "A": 24, "B": 15, "C": 42, "D": 13,
            "E": 18, "F": 18, "G": 55, "H": 28,
            "I": 55, "J": 55, "K": 25, "L": 55,
        }
        for column, width in widths.items():
            worksheet.column_dimensions[column].width = width

        from copy import copy
        from openpyxl.styles import Alignment, Font, PatternFill

        header_fill = PatternFill("solid", fgColor="1F4E78")
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    print("\nFinished.")
    print(f"Supplemental candidates saved to:\n{output_path}")


if __name__ == "__main__":
    main()
