from __future__ import annotations

import re
import time
from collections import defaultdict
from pathlib import Path
from typing import Any

import pandas as pd
from Bio import Entrez
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


START_DATE = "2025/07/01"
END_DATE = "2026/06/30"

ROSTER_FILE = "Penn_State_EM_Resident_Roster.xlsx"
OUTPUT_FILE = "FY26_Resident_PubMed_Review.xlsx"

# NCBI asks E-utilities users to provide an email address.
Entrez.email = "jeffrey.s.lubin@gmail.com"


PENN_STATE_AFFILIATION_QUERY = (
    '("Penn State"[Affiliation] '
    'OR "Pennsylvania State University"[Affiliation] '
    'OR "Penn State Health"[Affiliation] '
    'OR Hershey[Affiliation])'
)


def clean(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return re.sub(r"\s+", " ", str(value)).strip()


def split_variants(value: Any) -> list[str]:
    return [
        clean(part)
        for part in clean(value).split(";")
        if clean(part)
    ]


def resident_display_name(row: pd.Series) -> str:
    return clean(row.get("Full Name"))


def exact_full_name_terms(row: pd.Series) -> list[str]:
    """
    Build exact full-name searches.

    Most residents use First + Last. S. Michael Stuart is deliberately searched
    under both legal and preferred forms.
    """
    first = clean(row.get("First Name"))
    last = clean(row.get("Last Name"))
    display = clean(row.get("Full Name"))

    names: list[str] = []

    if first and last:
        names.append(f"{first} {last}")

    # Special handling for display names containing a preferred first name.
    if display == "S. Michael Stuart":
        names.extend(["Michael Stuart", "Shawn Stuart"])

    return list(dict.fromkeys(names))


def build_query(row: pd.Series) -> str:
    """
    High-sensitivity without uncontrolled common-name noise:

    1. Exact full-name searches are allowed even if affiliation is absent.
    2. Broad surname/initial searches are allowed only when a Penn State/Hershey
       affiliation is present somewhere in the PubMed record.
    3. Date basis is PubMed Entry Date (EDAT).
    """
    exact_terms = [
        f'"{name}"[Full Author Name]'
        for name in exact_full_name_terms(row)
    ]

    variant_terms = [
        f"{variant}[Author]"
        for variant in split_variants(row.get("Suggested Search Variants"))
    ]

    if not variant_terms:
        last = clean(row.get("Last Name"))
        initials = clean(row.get("PubMed Initials"))
        if last and initials:
            variant_terms.append(f"{last} {initials}[Author]")

    routes: list[str] = []

    if exact_terms:
        routes.append("(" + " OR ".join(exact_terms) + ")")

    if variant_terms:
        routes.append(
            "((" + " OR ".join(variant_terms) + ") "
            f"AND {PENN_STATE_AFFILIATION_QUERY})"
        )

    if not routes:
        raise ValueError(
            f"No usable search identity for {resident_display_name(row)}"
        )

    return (
        "(" + " OR ".join(routes) + ") "
        f"AND {START_DATE}:{END_DATE}[EDAT]"
    )


def search_pmids(query: str) -> list[str]:
    with Entrez.esearch(
        db="pubmed",
        term=query,
        retmax=10000,
        sort="pub date",
    ) as handle:
        result = Entrez.read(handle)
    return [str(pmid) for pmid in result.get("IdList", [])]


def fetch_records(pmids: list[str]) -> list[dict]:
    records: list[dict] = []

    for start in range(0, len(pmids), 200):
        batch = pmids[start:start + 200]

        with Entrez.efetch(
            db="pubmed",
            id=",".join(batch),
            rettype="medline",
            retmode="xml",
        ) as handle:
            result = Entrez.read(handle)

        records.extend(result.get("PubmedArticle", []))
        time.sleep(0.4)

    return records


def history_date(record: dict, wanted_status: str) -> str:
    for item in record.get("PubmedData", {}).get("History", []):
        status = clean(
            getattr(item, "attributes", {}).get("PubStatus")
        ).lower()

        if status != wanted_status.lower():
            continue

        year = clean(item.get("Year"))
        month = clean(item.get("Month")).zfill(2)
        day = clean(item.get("Day")).zfill(2)

        if year and month and day:
            return f"{year}-{month}-{day}"

    return ""


def pubmed_entry_date(record: dict) -> str:
    return (
        history_date(record, "entrez")
        or history_date(record, "pubmed")
    )


def publication_date(record: dict) -> str:
    pub_date = (
        record.get("MedlineCitation", {})
        .get("Article", {})
        .get("Journal", {})
        .get("JournalIssue", {})
        .get("PubDate", {})
    )

    year = clean(pub_date.get("Year"))
    month = clean(pub_date.get("Month"))
    day = clean(pub_date.get("Day"))
    medline_date = clean(pub_date.get("MedlineDate"))

    if year:
        return "-".join(part for part in [year, month, day] if part)

    return medline_date


def parse_record(record: dict) -> dict[str, Any]:
    citation = record.get("MedlineCitation", {})
    article = citation.get("Article", {})
    journal = article.get("Journal", {})

    author_details: list[dict[str, str]] = []
    authors: list[str] = []
    affiliations: list[str] = []

    for author in article.get("AuthorList", []):
        collective = clean(author.get("CollectiveName"))

        if collective:
            authors.append(collective)
            continue

        last = clean(author.get("LastName"))
        fore = clean(author.get("ForeName"))
        initials = clean(author.get("Initials"))

        author_details.append(
            {
                "last": last,
                "fore": fore,
                "initials": initials,
            }
        )

        authors.append(", ".join(part for part in [last, fore] if part))

        for affiliation_info in author.get("AffiliationInfo", []):
            affiliation = clean(affiliation_info.get("Affiliation"))
            if affiliation:
                affiliations.append(affiliation)

    doi = ""
    for article_id in record.get("PubmedData", {}).get("ArticleIdList", []):
        attributes = getattr(article_id, "attributes", {})
        if attributes.get("IdType") == "doi":
            doi = clean(article_id)
            break

    publication_types = [
        clean(item)
        for item in article.get("PublicationTypeList", [])
        if clean(item)
    ]

    return {
        "PMID": clean(citation.get("PMID")),
        "EDAT": pubmed_entry_date(record),
        "Publication Date": publication_date(record),
        "Title": clean(article.get("ArticleTitle")),
        "Journal": clean(journal.get("Title")),
        "Authors": "; ".join(authors),
        "Affiliations": " | ".join(dict.fromkeys(affiliations)),
        "DOI": doi,
        "Publication Types": "; ".join(publication_types),
        "_author_details": author_details,
    }


def has_penn_state_affiliation(affiliations: str) -> bool:
    text = affiliations.lower()

    return any(
        phrase in text
        for phrase in [
            "penn state",
            "pennsylvania state university",
            "penn state health",
            "hershey medical center",
            "hershey, pennsylvania",
            "hershey, pa",
        ]
    )


def exact_name_matches(row: pd.Series, article: dict[str, Any]) -> bool:
    target_last = clean(row.get("Last Name")).lower()
    expected_first_names = {
        name.split()[0].lower()
        for name in exact_full_name_terms(row)
        if name
    }

    for author in article["_author_details"]:
        if author["last"].lower() != target_last:
            continue

        fore = author["fore"].lower()

        for expected_first in expected_first_names:
            if (
                fore == expected_first
                or fore.startswith(expected_first + " ")
                or expected_first.startswith(fore + " ")
            ):
                return True

    return False


def initials_match(row: pd.Series, article: dict[str, Any]) -> bool:
    last = clean(row.get("Last Name")).lower()
    acceptable_initials = {
        re.sub(r"[^A-Za-z]", "", variant.split()[-1]).lower()
        for variant in split_variants(row.get("Suggested Search Variants"))
        if len(variant.split()) >= 2
    }

    for author in article["_author_details"]:
        if author["last"].lower() != last:
            continue

        article_initials = author["initials"].lower()

        for candidate in acceptable_initials:
            if article_initials.startswith(candidate) or candidate.startswith(article_initials):
                return True

    return False


def cohort_coauthors(
    row: pd.Series,
    article: dict[str, Any],
    roster: pd.DataFrame,
) -> list[str]:
    current_name = resident_display_name(row)
    article_people = {
        (
            author["last"].lower(),
            author["fore"].split()[0].lower() if author["fore"] else "",
        )
        for author in article["_author_details"]
    }

    matches: list[str] = []

    for _, other in roster.iterrows():
        other_name = resident_display_name(other)
        if not other_name or other_name == current_name:
            continue

        key = (
            clean(other.get("Last Name")).lower(),
            clean(other.get("First Name")).lower(),
        )

        if key in article_people:
            matches.append(other_name)

    return matches


def classify(
    row: pd.Series,
    article: dict[str, Any],
    roster: pd.DataFrame,
) -> tuple[str, str]:
    exact = exact_name_matches(row, article)
    initials = initials_match(row, article)
    penn_state = has_penn_state_affiliation(article["Affiliations"])
    coauthors = cohort_coauthors(row, article, roster)

    reasons: list[str] = []

    if exact:
        reasons.append("Exact full-name match")
    elif initials:
        reasons.append("Surname/initial match")

    if penn_state:
        reasons.append("Penn State/Hershey affiliation")

    if coauthors:
        reasons.append("Resident coauthor(s): " + ", ".join(coauthors))

    if exact and penn_state:
        return "Likely", "; ".join(reasons)

    if initials and penn_state and coauthors:
        return "Likely", "; ".join(reasons)

    if exact:
        return "Needs Review", "; ".join(reasons)

    if initials and penn_state:
        return "Needs Review", "; ".join(reasons)

    return "Unlikely", "; ".join(reasons) or "Insufficient identity support"


def shorten_for_excel(value: Any, limit: int = 32000) -> str:
    text = clean(value)

    if len(text) <= limit:
        return text

    return text[: limit - 24] + " ... [truncated]"


def add_review_sheet(
    workbook: Workbook,
    title: str,
    rows: list[dict[str, Any]],
) -> None:
    sheet = workbook.create_sheet(title)

    headers = [
        "Review Decision",
        "Resident Group",
        "Resident Name",
        "Confidence",
        "Why Included",
        "PMID",
        "EDAT",
        "Publication Date",
        "Title",
        "Journal",
        "Authors",
        "Affiliations",
        "DOI",
        "Publication Types",
        "Reviewer Notes",
    ]

    sheet.append(headers)

    for item in rows:
        sheet.append(
            [
                "",
                item["Resident Group"],
                item["Resident Name"],
                item["Confidence"],
                item["Why Included"],
                item["PMID"],
                item["EDAT"],
                item["Publication Date"],
                item["Title"],
                item["Journal"],
                item["Authors"],
                item["Affiliations"],
                item["DOI"],
                item["Publication Types"],
                "",
            ]
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    widths = {
        "A": 18,
        "B": 18,
        "C": 24,
        "D": 15,
        "E": 45,
        "F": 13,
        "G": 14,
        "H": 17,
        "I": 55,
        "J": 30,
        "K": 55,
        "L": 55,
        "M": 25,
        "N": 25,
        "O": 30,
    }

    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    for row_cells in sheet.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    decision_validation = DataValidation(
        type="list",
        formula1='"Keep,Exclude,Unsure"',
        allow_blank=True,
    )
    sheet.add_data_validation(decision_validation)
    decision_validation.add(f"A2:A{max(sheet.max_row, 2)}")

    green = PatternFill("solid", fgColor="E2F0D9")
    red = PatternFill("solid", fgColor="F4CCCC")
    yellow = PatternFill("solid", fgColor="FFF2CC")

    sheet.conditional_formatting.add(
        f"A2:A{max(sheet.max_row, 2)}",
        FormulaRule(formula=['A2="Keep"'], fill=green),
    )
    sheet.conditional_formatting.add(
        f"A2:A{max(sheet.max_row, 2)}",
        FormulaRule(formula=['A2="Exclude"'], fill=red),
    )
    sheet.conditional_formatting.add(
        f"A2:A{max(sheet.max_row, 2)}",
        FormulaRule(formula=['A2="Unsure"'], fill=yellow),
    )


def main() -> None:
    folder = Path(__file__).resolve().parent
    roster_path = folder / ROSTER_FILE
    output_path = folder / OUTPUT_FILE

    if not roster_path.exists():
        raise FileNotFoundError(
            f"Could not find {ROSTER_FILE} in {folder}."
        )

    roster = pd.read_excel(
        roster_path,
        sheet_name="Resident Roster",
    )

    required_columns = {
        "Resident Group",
        "Full Name",
        "First Name",
        "Last Name",
        "Suggested Search Variants",
    }

    missing = required_columns - set(roster.columns)

    if missing:
        raise ValueError(
            "Resident Roster is missing: "
            + ", ".join(sorted(missing))
        )

    roster = roster.dropna(subset=["Full Name", "Last Name"]).copy()

    all_rows: list[dict[str, Any]] = []
    article_cache: dict[str, dict[str, Any]] = {}

    print(f"Searching {len(roster)} residents...")
    print(f"PubMed Entry Date range: {START_DATE} through {END_DATE}\n")

    for number, (_, resident) in enumerate(roster.iterrows(), start=1):
        name = resident_display_name(resident)
        query = build_query(resident)

        print(f"[{number}/{len(roster)}] {name}")
        pmids = search_pmids(query)
        print(f"    Candidate PubMed records: {len(pmids)}")

        missing_pmids = [
            pmid
            for pmid in pmids
            if pmid not in article_cache
        ]

        if missing_pmids:
            for raw_record in fetch_records(missing_pmids):
                article = parse_record(raw_record)

                if article["PMID"]:
                    article_cache[article["PMID"]] = article

        for pmid in pmids:
            article = article_cache.get(pmid)

            if not article:
                continue

            confidence, reason = classify(
                resident,
                article,
                roster,
            )

            all_rows.append(
                {
                    "Resident Group": clean(resident.get("Resident Group")),
                    "Resident Name": name,
                    "Confidence": confidence,
                    "Why Included": reason,
                    "PMID": article["PMID"],
                    "EDAT": article["EDAT"],
                    "Publication Date": article["Publication Date"],
                    "Title": article["Title"],
                    "Journal": article["Journal"],
                    "Authors": article["Authors"],
                    "Affiliations": shorten_for_excel(article["Affiliations"]),
                    "DOI": article["DOI"],
                    "Publication Types": article["Publication Types"],
                }
            )

        time.sleep(0.4)

    # Remove duplicate resident-PMID matches.
    unique_rows: dict[tuple[str, str], dict[str, Any]] = {}

    for row in all_rows:
        key = (row["Resident Name"], row["PMID"])
        unique_rows[key] = row

    rows = list(unique_rows.values())

    rank = {
        "Likely": 0,
        "Needs Review": 1,
        "Unlikely": 2,
    }

    rows.sort(
        key=lambda item: (
            rank.get(item["Confidence"], 9),
            item["Resident Group"],
            item["Resident Name"],
            item["EDAT"],
            item["PMID"],
        )
    )

    likely = [
        row for row in rows
        if row["Confidence"] == "Likely"
    ]

    needs_review = [
        row for row in rows
        if row["Confidence"] == "Needs Review"
    ]

    unlikely = [
        row for row in rows
        if row["Confidence"] == "Unlikely"
    ]

    workbook = Workbook()

    add_review_sheet(
        workbook,
        "Likely",
        likely,
    )

    add_review_sheet(
        workbook,
        "Needs Review",
        needs_review,
    )

    # Keep excluded candidates available for audit, but do not make them part
    # of the primary review burden.
    add_review_sheet(
        workbook,
        "Excluded",
        unlikely,
    )

    summary = workbook.create_sheet("Summary", 0)
    summary.append(
        [
            "Resident Name",
            "Resident Group",
            "Likely",
            "Needs Review",
            "Excluded",
            "Total Retrieved",
        ]
    )

    counts: defaultdict[str, dict[str, int]] = defaultdict(
        lambda: {
            "Likely": 0,
            "Needs Review": 0,
            "Unlikely": 0,
        }
    )

    group_by_name: dict[str, str] = {}

    for row in rows:
        name = row["Resident Name"]
        counts[name][row["Confidence"]] += 1
        group_by_name[name] = row["Resident Group"]

    for _, resident in roster.iterrows():
        name = resident_display_name(resident)
        data = counts[name]

        summary.append(
            [
                name,
                clean(resident.get("Resident Group")),
                data["Likely"],
                data["Needs Review"],
                data["Unlikely"],
                data["Likely"] + data["Needs Review"] + data["Unlikely"],
            ]
        )

    summary.freeze_panes = "A2"
    summary.auto_filter.ref = summary.dimensions

    for cell in summary[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for column, width in {
        "A": 24,
        "B": 18,
        "C": 12,
        "D": 16,
        "E": 12,
        "F": 16,
    }.items():
        summary.column_dimensions[column].width = width

    workbook.save(output_path)

    print("\nFinished.")
    print(f"Likely records: {len(likely)}")
    print(f"Needs Review records: {len(needs_review)}")
    print(f"Excluded records: {len(unlikely)}")
    print(f"Output saved to:\n{output_path}")


if __name__ == "__main__":
    main()
