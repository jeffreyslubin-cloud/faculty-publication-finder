from __future__ import annotations

import re
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from Bio import Entrez

INPUT_FILE = "Penn_State_EM_Faculty.xlsx"
Entrez.email = "jeffrey.s.lubin@gmail.com"


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return re.sub(r"\s+", " ", str(value)).strip()


def prompt_date(label: str) -> str:
    formats = ("%m/%d/%Y", "%Y/%m/%d", "%Y-%m-%d", "%m-%d-%Y")
    while True:
        raw = input(f"{label} (MM/DD/YYYY): ").strip()
        for date_format in formats:
            try:
                return datetime.strptime(raw, date_format).strftime("%Y/%m/%d")
            except ValueError:
                continue
        print("Invalid date. Example: 07/01/2026")


def canonical_faculty_name(row: pd.Series) -> str:
    last = clean_text(row.get("Last Name"))
    first = clean_text(row.get("First Name / Initial"))
    return f"{last}, {first}" if first else last


def build_query(row: pd.Series, start_date: str, end_date: str) -> str:
    last = clean_text(row.get("Last Name"))
    first = clean_text(row.get("First Name / Initial"))
    initials = clean_text(row.get("PubMed Initials"))
    if not last or not initials:
        raise ValueError(f"Missing Last Name or PubMed Initials for {canonical_faculty_name(row)}")

    terms: list[str] = []
    if len(first) > 1:
        terms.append(f'"{first} {last}"[Full Author Name]')
    terms.append(f'"{last} {initials}"[Author]')

    return f'({" OR ".join(terms)}) AND ("{start_date}"[EDAT] : "{end_date}"[EDAT])'


def search_pmids(query: str) -> list[str]:
    with Entrez.esearch(db="pubmed", term=query, retmax=10000, sort="pub date") as handle:
        result = Entrez.read(handle)
    return [str(value) for value in result.get("IdList", [])]


def fetch_records(pmids: list[str]) -> list[dict]:
    records: list[dict] = []
    for start in range(0, len(pmids), 200):
        batch = pmids[start:start + 200]
        with Entrez.efetch(db="pubmed", id=",".join(batch), rettype="medline", retmode="xml") as handle:
            data = Entrez.read(handle)
        records.extend(data.get("PubmedArticle", []))
        time.sleep(0.4)
    return records


def history_date(record: dict, wanted_status: str) -> str:
    for item in record.get("PubmedData", {}).get("History", []):
        status = clean_text(getattr(item, "attributes", {}).get("PubStatus")).lower()
        if status == wanted_status.lower():
            year = clean_text(item.get("Year"))
            month = clean_text(item.get("Month")).zfill(2)
            day = clean_text(item.get("Day")).zfill(2)
            return "-".join(part for part in [year, month, day] if part)
    return ""


def article_date(record: dict) -> str:
    pub_date = (
        record.get("MedlineCitation", {})
        .get("Article", {})
        .get("Journal", {})
        .get("JournalIssue", {})
        .get("PubDate", {})
    )
    year = clean_text(pub_date.get("Year"))
    month = clean_text(pub_date.get("Month"))
    day = clean_text(pub_date.get("Day"))
    medline_date = clean_text(pub_date.get("MedlineDate"))
    if year:
        return "-".join(part for part in [year, month, day] if part)
    return medline_date


def parse_article(record: dict) -> dict[str, Any]:
    citation = record.get("MedlineCitation", {})
    article = citation.get("Article", {})
    journal = article.get("Journal", {})

    authors: list[str] = []
    author_details: list[dict[str, str]] = []
    affiliations: list[str] = []

    for author in article.get("AuthorList", []):
        collective = clean_text(author.get("CollectiveName"))
        if collective:
            authors.append(collective)
            continue

        last = clean_text(author.get("LastName"))
        fore = clean_text(author.get("ForeName"))
        initials = clean_text(author.get("Initials"))
        authors.append(", ".join(part for part in [last, fore] if part))
        author_details.append({"last": last, "fore": fore, "initials": initials})

        for affiliation_info in author.get("AffiliationInfo", []):
            affiliation = clean_text(affiliation_info.get("Affiliation"))
            if affiliation:
                affiliations.append(affiliation)

    doi = ""
    for article_id in record.get("PubmedData", {}).get("ArticleIdList", []):
        if getattr(article_id, "attributes", {}).get("IdType") == "doi":
            doi = clean_text(article_id)
            break

    publication_types = [
        clean_text(item)
        for item in article.get("PublicationTypeList", [])
        if clean_text(item)
    ]

    return {
        "PMID": clean_text(citation.get("PMID")),
        "Title": clean_text(article.get("ArticleTitle")),
        "Journal": clean_text(journal.get("Title")),
        "PubMed Entry Date": history_date(record, "entrez") or history_date(record, "pubmed"),
        "Publication Date": article_date(record),
        "Authors": "; ".join(authors),
        "Affiliations": " | ".join(dict.fromkeys(affiliations)),
        "DOI": doi,
        "Publication Types": "; ".join(publication_types),
        "_author_details": author_details,
    }


def affiliation_flags(row: pd.Series, affiliation_text: str) -> tuple[bool, bool, bool]:
    text = affiliation_text.lower()
    institution_match = any(x in text for x in ["penn state", "pennsylvania state", "hershey"])
    em_match = any(x in text for x in ["emergency medicine", "department of emergency", "dept of emergency"])

    secondary_match = False
    secondary = clean_text(row.get("Secondary Institution(s)"))
    for part in re.split(r"[;,|]", secondary):
        part = part.strip().lower()
        if part and part in text:
            secondary_match = True
            break

    return institution_match, em_match, secondary_match


def classify_match(row: pd.Series, article: dict[str, Any]) -> tuple[str, str, bool, bool]:
    last = clean_text(row.get("Last Name")).lower()
    first = clean_text(row.get("First Name / Initial")).lower()
    initials = clean_text(row.get("PubMed Initials")).lower()

    surname_matches = [a for a in article["_author_details"] if a["last"].lower() == last]
    institution_match, em_match, secondary_match = affiliation_flags(row, article["Affiliations"])

    if not surname_matches:
        return "Unlikely", "No exact surname match in parsed author list", institution_match, em_match

    if len(first) > 1:
        for author in surname_matches:
            fore = author["fore"].lower()
            full_name_match = fore == first or fore.startswith(first + " ") or first.startswith(fore + " ")
            if full_name_match:
                reasons = ["Exact given-name and surname match"]
                if institution_match:
                    reasons.append("Penn State/Hershey affiliation")
                if em_match:
                    reasons.append("Emergency Medicine affiliation")
                elif secondary_match:
                    reasons.append("secondary affiliation")
                return "Likely", "; ".join(reasons), institution_match, em_match

        for author in surname_matches:
            if author["initials"].lower().startswith(initials):
                if institution_match and em_match:
                    return "Likely", "Surname/initials match; Penn State/Hershey and Emergency Medicine affiliations", institution_match, em_match
                if secondary_match:
                    return "Likely", "Surname/initials match; secondary affiliation", institution_match, em_match
                return "Needs Review", "Surname and initials match, but affiliation support is incomplete", institution_match, em_match

        return "Unlikely", "Surname matches, but given name and initials do not", institution_match, em_match

    for author in surname_matches:
        if author["initials"].lower().startswith(initials):
            if institution_match and em_match:
                return "Likely", "Initial/surname match; Penn State/Hershey and Emergency Medicine affiliations", institution_match, em_match
            if secondary_match:
                return "Likely", "Initial/surname match; secondary affiliation", institution_match, em_match
            return "Needs Review", "Roster contains only an initial and affiliation support is incomplete", institution_match, em_match

    return "Unlikely", "Surname matches, but initials do not", institution_match, em_match


def keep_candidate(confidence: str, institution_match: bool, em_match: bool) -> bool:
    if confidence == "Likely":
        return True
    if confidence == "Needs Review":
        return institution_match and em_match
    return False


def format_sheet(writer: pd.ExcelWriter, sheet_name: str, dataframe: pd.DataFrame) -> None:
    worksheet = writer.book[sheet_name]
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_number, column_name in enumerate(dataframe.columns, start=1):
        values = [clean_text(column_name)]
        values.extend(clean_text(value) for value in dataframe[column_name].head(300))
        width = min(max(len(value) for value in values) + 2, 60)
        letter = worksheet.cell(1, column_number).column_letter
        worksheet.column_dimensions[letter].width = width

    from openpyxl.styles import Alignment, Font, PatternFill
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def main() -> None:
    folder = Path(__file__).resolve().parent
    input_path = folder / INPUT_FILE

    if not input_path.exists():
        raise FileNotFoundError(
            f"Could not find {INPUT_FILE} in {folder}. Place the workbook and script in the same folder."
        )

    print("Penn State Emergency Medicine PubMed search")
    print("Search dates are based on PubMed Entry Date (EDAT).\n")

    start_date = prompt_date("Start date")
    end_date = prompt_date("End date")

    if datetime.strptime(start_date, "%Y/%m/%d") > datetime.strptime(end_date, "%Y/%m/%d"):
        raise ValueError("The start date cannot be after the end date.")

    output_name = (
        "Penn_State_EM_EDAT_Results_"
        f"{start_date.replace('/', '')}_to_{end_date.replace('/', '')}.xlsx"
    )
    output_path = folder / output_name

    roster = pd.read_excel(input_path, sheet_name="Faculty Roster")
    required_columns = {"Last Name", "First Name / Initial", "PubMed Initials"}
    missing_columns = required_columns - set(roster.columns)
    if missing_columns:
        raise ValueError("The Faculty Roster sheet is missing: " + ", ".join(sorted(missing_columns)))

    roster = roster.dropna(subset=["Last Name", "PubMed Initials"]).copy()
    kept_matches: list[dict[str, Any]] = []
    excluded_counts: defaultdict[str, int] = defaultdict(int)
    article_cache: dict[str, dict[str, Any]] = {}

    print(f"\nSearching {len(roster)} faculty members...")
    print(f"EDAT range: {start_date} through {end_date}\n")

    for number, (_, faculty) in enumerate(roster.iterrows(), start=1):
        faculty_name = canonical_faculty_name(faculty)
        query = build_query(faculty, start_date, end_date)
        print(f"[{number}/{len(roster)}] {faculty_name}")
        pmids = search_pmids(query)
        print(f"    Raw PubMed results: {len(pmids)}")

        missing_pmids = [pmid for pmid in pmids if pmid not in article_cache]
        if missing_pmids:
            for raw_record in fetch_records(missing_pmids):
                parsed = parse_article(raw_record)
                if parsed["PMID"]:
                    article_cache[parsed["PMID"]] = parsed

        kept_for_person = 0
        for pmid in pmids:
            article = article_cache.get(pmid)
            if not article:
                continue

            confidence, reason, institution_match, em_match = classify_match(faculty, article)
            if not keep_candidate(confidence, institution_match, em_match):
                excluded_counts[confidence] += 1
                continue

            kept_for_person += 1
            kept_matches.append({
                "Faculty Name": faculty_name,
                "Confidence": confidence,
                "Reason": reason,
                "PMID": article["PMID"],
                "PubMed Entry Date": article["PubMed Entry Date"],
                "Publication Date": article["Publication Date"],
                "Title": article["Title"],
                "Journal": article["Journal"],
                "Authors": article["Authors"],
                "Affiliations": article["Affiliations"],
                "Penn State/Hershey Affiliation": "Yes" if institution_match else "No",
                "Emergency Medicine Affiliation": "Yes" if em_match else "No",
                "DOI": article["DOI"],
                "Publication Types": article["Publication Types"],
                "Search Query": query,
            })

        print(f"    Candidates retained: {kept_for_person}")
        time.sleep(0.4)

    columns = [
        "Faculty Name", "Confidence", "Reason", "PMID", "PubMed Entry Date",
        "Publication Date", "Title", "Journal", "Authors", "Affiliations",
        "Penn State/Hershey Affiliation", "Emergency Medicine Affiliation",
        "DOI", "Publication Types", "Search Query",
    ]
    results = pd.DataFrame(kept_matches, columns=columns)

    if not results.empty:
        rank = {"Likely": 0, "Needs Review": 1}
        results["_rank"] = results["Confidence"].map(rank).fillna(9)
        results = (
            results.sort_values(["_rank", "Faculty Name", "PubMed Entry Date", "PMID"], ascending=[True, True, False, True])
            .drop(columns=["_rank"])
            .drop_duplicates(subset=["Faculty Name", "PMID"])
        )

    unique_columns = [
        "PMID", "PubMed Entry Date", "Publication Date", "Title", "Journal",
        "Authors", "Affiliations", "DOI", "Publication Types", "Matched Faculty",
    ]

    if results.empty:
        unique = pd.DataFrame(columns=unique_columns)
        summary = pd.DataFrame(columns=["Faculty Name", "Likely", "Needs Review", "Total Retained"])
    else:
        unique = (
            results.groupby("PMID", as_index=False)
            .agg({
                "PubMed Entry Date": "first",
                "Publication Date": "first",
                "Title": "first",
                "Journal": "first",
                "Authors": "first",
                "Affiliations": "first",
                "DOI": "first",
                "Publication Types": "first",
                "Faculty Name": lambda values: "; ".join(dict.fromkeys(values)),
            })
            .rename(columns={"Faculty Name": "Matched Faculty"})
        )

        summary = (
            results.groupby(["Faculty Name", "Confidence"], dropna=False)
            .size()
            .unstack(fill_value=0)
            .reset_index()
        )
        for column in ["Likely", "Needs Review"]:
            if column not in summary.columns:
                summary[column] = 0
        summary["Total Retained"] = summary["Likely"] + summary["Needs Review"]
        summary = summary[["Faculty Name", "Likely", "Needs Review", "Total Retained"]]

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        unique.to_excel(writer, sheet_name="Unique Publications", index=False)
        results.to_excel(writer, sheet_name="Faculty Matches", index=False)
        summary.to_excel(writer, sheet_name="Review Summary", index=False)
        format_sheet(writer, "Unique Publications", unique)
        format_sheet(writer, "Faculty Matches", results)
        format_sheet(writer, "Review Summary", summary)

    print("\nFinished.")
    print(f"Retained faculty-publication matches: {len(results)}")
    print(f"Unique publications retained: {len(unique)}")
    if excluded_counts:
        print("Excluded matches: " + ", ".join(f"{k}={v}" for k, v in sorted(excluded_counts.items())))
    print(f"\nResults saved to:\n{output_path}")


if __name__ == "__main__":
    main()
