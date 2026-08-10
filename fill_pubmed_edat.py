from __future__ import annotations

import time
from pathlib import Path

from Bio import Entrez
from openpyxl import load_workbook


INPUT_FILE = "FY26_Combined_Publications.xlsx"
OUTPUT_FILE = "FY26_Combined_Publications_With_EDAT.xlsx"

Entrez.email = "jeffrey.s.lubin@gmail.com"


def clean(value) -> str:
    return "" if value is None else str(value).strip()


def pubmed_entry_date(article: dict) -> str:
    history = article.get("PubmedData", {}).get("History", [])

    for wanted_status in ("entrez", "pubmed"):
        for item in history:
            status = clean(getattr(item, "attributes", {}).get("PubStatus")).lower()
            if status != wanted_status:
                continue

            year = clean(item.get("Year"))
            month = clean(item.get("Month")).zfill(2)
            day = clean(item.get("Day")).zfill(2)

            if year and month and day:
                return f"{year}-{month}-{day}"

    return ""


def fetch_edats(pmids: list[str]) -> dict[str, str]:
    result: dict[str, str] = {}

    for start in range(0, len(pmids), 200):
        batch = pmids[start:start + 200]

        with Entrez.efetch(
            db="pubmed",
            id=",".join(batch),
            rettype="medline",
            retmode="xml",
        ) as handle:
            data = Entrez.read(handle)

        for article in data.get("PubmedArticle", []):
            citation = article.get("MedlineCitation", {})
            pmid = clean(citation.get("PMID"))
            if pmid:
                result[pmid] = pubmed_entry_date(article)

        time.sleep(0.4)

    return result


def main() -> None:
    folder = Path(__file__).resolve().parent
    input_path = folder / INPUT_FILE
    output_path = folder / OUTPUT_FILE

    if not input_path.exists():
        raise FileNotFoundError(
            f"Could not find {INPUT_FILE} in {folder}."
        )

    workbook = load_workbook(input_path)
    sheet = workbook["Combined Publications"]

    headers = {
        clean(cell.value): cell.column
        for cell in sheet[1]
        if clean(cell.value)
    }

    required = {"PMID", "EDAT"}
    missing = required - set(headers)
    if missing:
        raise ValueError(
            "Missing required column(s): " + ", ".join(sorted(missing))
        )

    pmid_col = headers["PMID"]
    edat_col = headers["EDAT"]

    rows_by_pmid: dict[str, list[int]] = {}

    for row_number in range(2, sheet.max_row + 1):
        pmid = clean(sheet.cell(row=row_number, column=pmid_col).value)
        if pmid:
            rows_by_pmid.setdefault(pmid, []).append(row_number)

    pmids = list(rows_by_pmid)
    print(f"Fetching EDAT values for {len(pmids)} unique PMIDs...")

    edats = fetch_edats(pmids)

    filled = 0
    missing_edat = []

    for pmid, row_numbers in rows_by_pmid.items():
        edat = edats.get(pmid, "")

        if not edat:
            missing_edat.append(pmid)
            continue

        for row_number in row_numbers:
            cell = sheet.cell(row=row_number, column=edat_col)
            cell.value = edat
            cell.number_format = "yyyy-mm-dd"
            filled += 1

    workbook.save(output_path)

    print(f"EDAT cells filled: {filled}")
    print(f"PMIDs with no EDAT returned: {len(missing_edat)}")

    if missing_edat:
        print("No EDAT returned for:")
        for pmid in missing_edat:
            print(f"  {pmid}")

    print("\nFinished.")
    print(f"Corrected workbook saved to:\n{output_path}")


if __name__ == "__main__":
    main()
